import json
from pymongo import MongoClient
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Configuración
MONGO_URI = 'localhost:27017'
DB_NAME = 'api'
SCHEMA_PATH = 'C:\Proyecto-VCP\generador-nombres-clikealo\schemaLaptop.json'

# Colores para cada pestaña (en formato RGB hexadecimal)
COLORES_PESTAÑAS = {
    "productos originales": "4472C4",  # Azul
    "productos despues de abasteo": "ED7D31",  # Naranja
    "productos despues de icecat": "A5A5A5",  # Gris
    "productos completos": "70AD47"  # Verde
}

# Cargar el schema
with open(SCHEMA_PATH, 'r', encoding='utf-8') as f:
    # Remover comentarios del JSON
    content = f.read()
    lines = content.split('\n')
    clean_lines = [line.split('//')[0] for line in lines]
    clean_content = '\n'.join(clean_lines)
    schema = json.loads(clean_content)

# Función para convertir plural a singular en español
def plural_a_singular(palabra):
    # Convertir una palabra de plural a singular
    palabra_lower = palabra.lower()
    
    # Reglas de pluralización en español (aplicadas en reversa)
    # Palabras que terminan en -ces → -z (ej: luces → luz)
    if palabra_lower.endswith('ces'):
        base = palabra[:-3]
        return base + 'z'
    
    # Palabras que terminan en -es después de consonante (ej: cables → cable)
    elif palabra_lower.endswith('es') and len(palabra) > 2:
        # Verificar si antes de 'es' hay una consonante
        if palabra[-3] not in 'aeiouáéíóú':
            return palabra[:-2]
        # Si termina en -ies, cambiar a -y (aunque es menos común en español)
        elif palabra_lower.endswith('ies'):
            return palabra[:-3] + 'y'
        # Para palabras que terminan en vocal + es, quitar solo la 's'
        else:
            return palabra[:-1]
    
    # Palabras que terminan en -s después de vocal (ej: laptops → laptop, autos → auto)
    elif palabra_lower.endswith('s') and len(palabra) > 1:
        # Verificar si antes de 's' hay una vocal
        if palabra[-2] in 'aeiouáéíóú':
            return palabra[:-1]
        # Si termina en consonante + s, generalmente quitar solo la 's'
        else:
            return palabra[:-1]
    
    # Si no coincide con ninguna regla, devolver la palabra original
    return palabra

# Función para convertir texto de plural a singular
def texto_a_singular(texto):
    if not texto:
        return texto
    
    # Dividir el texto en palabras
    palabras = texto.split()
    palabras_singulares = []
    
    for palabra in palabras:
        # Preservar mayúsculas/minúsculas originales
        if palabra.isupper():
            # Si toda la palabra está en mayúsculas
            palabras_singulares.append(plural_a_singular(palabra).upper())
        elif palabra[0].isupper():
            # Si solo la primera letra está en mayúscula
            singular = plural_a_singular(palabra)
            palabras_singulares.append(singular[0].upper() + singular[1:])
        else:
            # Minúsculas
            palabras_singulares.append(plural_a_singular(palabra))
    
    return ' '.join(palabras_singulares)

# Función para aplicar transformación de texto
def aplicar_transformacion(texto, transformacion):
    if not texto:
        return None
    
    if transformacion == 'mayuscula':
        return texto.upper()
    elif transformacion == 'minuscula':
        return texto.lower()
    elif transformacion == 'capitalize':
        return texto.capitalize()
    elif transformacion == 'singular':
        return texto_a_singular(texto)
    else:  # 'ninguna'
        return texto

# Función para extraer valor de especificaciones
def extraer_especificacion(especificaciones, titulo_seccion, dato):
    for seccion in especificaciones:
        if seccion.get('tituloSeccion') == titulo_seccion:
            for item in seccion.get('seccionList', []):
                if item.get('dato') == dato:
                    return item.get('valor')
    return None

# Función para procesar un campo del schema
def procesar_campo(producto, campo_config):
    # Si es solo texto estático
    if 'texto' in campo_config:
        return {'valor': campo_config['texto'], 'faltante': False, 'esTexto': True}
    
    valor = None
    
    # Procesar según el tipo de campo
    if campo_config['campo'] == 'marca':
        valor = producto.get('marca')
    
    elif campo_config['campo'] == 'nombreProducto':
        valor = producto.get('nombreProducto')
    
    elif campo_config['campo'] == 'categorias':
        index = campo_config.get('index', 0)
        categorias = producto.get('categorias', [])
        if categorias:
            categoria = categorias[index] if index != -1 else categorias[-1]
            if categoria and 'subcampo' in campo_config:
                valor = categoria.get(campo_config['subcampo'])
    
    elif campo_config['campo'] == 'especificaciones' and 'condicion' in campo_config:
        titulo_seccion = campo_config['condicion']['tituloSeccion']
        dato = campo_config['subcampo']['seccionList']['condicion']['dato']
        valor = extraer_especificacion(producto.get('especificaciones', []), titulo_seccion, dato)
    
    # Aplicar transformación si existe
    if valor and 'transformacion' in campo_config:
        valor = aplicar_transformacion(valor, campo_config['transformacion'])
    
    return {
        'valor': valor,
        'faltante': not valor,
        'esTexto': False
    }

# Función para obtener el nombre del campo para el encabezado
def obtener_nombre_campo(campo_config):
    if campo_config['campo'] == 'especificaciones' and 'condicion' in campo_config:
        # Para especificaciones, usar el nombre del dato específico
        dato = campo_config['subcampo']['seccionList']['condicion']['dato']
        return dato
    else:
        # Para otros campos, usar el nombre del campo
        return campo_config['campo']

# Función para generar el nombre del producto
def generar_nombre_producto(producto, estructura):
    partes = []
    partes_detalle = []
    algun_faltante = False
    
    for campo_config in estructura:
        resultado = procesar_campo(producto, campo_config)
        
        # Para el nombre completo
        if resultado['faltante'] and 'campo' in campo_config:
            partes.append(f"[{campo_config['campo'].upper()}]")
            algun_faltante = True
        elif resultado['valor']:
            partes.append(resultado['valor'])
        
        # Para las columnas individuales (solo campos, no texto estático)
        if 'campo' in campo_config:
            nombre_campo = obtener_nombre_campo(campo_config)
            partes_detalle.append({
                'nombre': nombre_campo,
                'valor': f"[{campo_config['campo'].upper()}]" if resultado['faltante'] else resultado['valor'],
                'faltante': resultado['faltante']
            })
    
    return {
        'nombreCompleto': ' '.join(partes).strip(),
        'partes': partes_detalle,
        'algunFaltante': algun_faltante
    }

# Función auxiliar para escribir datos en una pestaña específica
def escribir_en_pestaña(productos, estructura, nombre_archivo, nombre_pestaña):

    generar_excel(productos, estructura, nombre_archivo, nombre_pestaña)

# Función para generar Excel
def generar_excel(productos, estructura, nombre_archivo, nombre_pestaña="Productos"):
    """
    Actualiza SOLO la pestaña especificada, sin modificar las demás pestañas del documento.
    
    Args:
        productos: Lista de productos a escribir
        estructura: Estructura del nombre del producto (del schema)
        nombre_archivo: Nombre del archivo Excel
        nombre_pestaña: Nombre de la pestaña a actualizar (las demás pestañas no se tocan)
    """
    # Cargar workbook existente (preserva todas las pestañas)
    try:
        wb = openpyxl.load_workbook(nombre_archivo)
        # Guardar lista de pestañas existentes para verificación
        pestañas_existentes = wb.sheetnames.copy()
    except FileNotFoundError:
        # Si el archivo no existe, crear uno nuevo
        wb = openpyxl.Workbook()
        # Eliminar la pestaña por defecto si es un nuevo workbook
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        pestañas_existentes = []
    
    # Seleccionar SOLO la pestaña indicada (las demás no se tocan)
    if nombre_pestaña in wb.sheetnames:
        ws = wb[nombre_pestaña]
        # Limpiar contenido existente solo de esta pestaña
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
    else:
        # Crear la pestaña si no existe (sin afectar las demás)
        ws = wb.create_sheet(nombre_pestaña)
    
    # Aplicar color a la pestaña si está definido
    if nombre_pestaña in COLORES_PESTAÑAS:
        ws.sheet_properties.tabColor = COLORES_PESTAÑAS[nombre_pestaña]
    
    # Definir estilos
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    verde_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    amarillo_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    celda_faltante_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
    celda_faltante_font = Font(bold=True)
    
    # Preparar encabezados
    headers = ['SKU', 'Nombre Completo']
    for campo in estructura:
        if 'campo' in campo:
            nombre_campo = obtener_nombre_campo(campo)
            headers.append(nombre_campo)
    
    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Escribir datos
    for row_num, producto in enumerate(productos, 2):
        resultado = generar_nombre_producto(producto, estructura)
        
        # Determinar color de fila
        row_fill = amarillo_fill if resultado['algunFaltante'] else verde_fill
        
        # SKU
        cell = ws.cell(row=row_num, column=1)
        cell.value = producto.get('sku', '')
        cell.fill = row_fill
        
        # Nombre completo
        cell = ws.cell(row=row_num, column=2)
        cell.value = resultado['nombreCompleto']
        cell.fill = row_fill
        
        # Partes individuales
        for col_num, parte in enumerate(resultado['partes'], 3):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = parte['valor'] or ''
            
            # Si la celda tiene dato faltante, aplicar estilo especial
            if parte['faltante']:
                cell.fill = celda_faltante_fill
                cell.font = celda_faltante_font
            else:
                cell.fill = row_fill
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar archivo (solo se actualiza la pestaña indicada, las demás se preservan intactas)
    wb.save(nombre_archivo)

# Función principal
def main():
    try:
        print('Conectando a MongoDB...')
        client = MongoClient(MONGO_URI)
        print('Conectado exitosamente')
        
        db = client[DB_NAME]
        collection = db[schema['coleccion']]
        
        # Buscar productos según el tipo
        query = {'categorias.clave': schema['tipo']}
        productos = list(collection.find(query))
        
        print(f'Productos encontrados: {len(productos)}')
        
        if len(productos) == 0:
            print('No se encontraron productos para procesar')
            return
        
        # Generar Excel con las cuatro pestañas
        nombre_archivo = 'productos_output.xlsx'
        
        # Crear las cuatro pestañas desde el inicio
        pestañas = [
            "productos originales",
            "productos despues de abasteo",
            "productos despues de icecat",
            "productos completos"
        ]
        
        # Crear o cargar workbook existente (preserva datos existentes)
        try:
            wb = openpyxl.load_workbook(nombre_archivo)
        except FileNotFoundError:
            # Si el archivo no existe, crear uno nuevo
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
        
        # Crear solo las pestañas que no existan (preserva las que ya tienen datos)
        for pestaña in pestañas:
            if pestaña not in wb.sheetnames:
                ws = wb.create_sheet(pestaña)
                # Aplicar color a la pestaña
                if pestaña in COLORES_PESTAÑAS:
                    ws.sheet_properties.tabColor = COLORES_PESTAÑAS[pestaña]
            else:
                # Si la pestaña ya existe, solo actualizar el color si es necesario
                ws = wb[pestaña]
                if pestaña in COLORES_PESTAÑAS:
                    ws.sheet_properties.tabColor = COLORES_PESTAÑAS[pestaña]
        
        # Guardar solo si se crearon nuevas pestañas o se modificaron colores
        wb.save(nombre_archivo)
        
        
        pestaña_destino = "productos originales"  # 
        
        # Escribir datos en la pestaña seleccionada
        escribir_en_pestaña(productos, schema['estructuraNombreProducto'], nombre_archivo, pestaña_destino)
        
        
        print(f'Excel generado: {nombre_archivo}')
        print(f'Total de productos procesados: {len(productos)}')
        print(f'Pestañas creadas: {", ".join(pestañas)}')
        print(f'Datos escritos en la pestaña: "{pestaña_destino}"')
        
    except Exception as error:
        print(f'Error: {error}')
    finally:
        client.close()
        print('Conexión cerrada')

# Ejecutar
if __name__ == '__main__':
    main()