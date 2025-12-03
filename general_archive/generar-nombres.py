import json
import os
import glob
from pymongo import MongoClient
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# Configuración
MONGO_URI = 'MONGO_URI PARA LA BDD EN PRODUCCION'
DB_NAME = 'development'

# Obtener el directorio del script y construir la ruta a schemas
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SCHEMAS_DIR = os.path.join(SCRIPT_DIR, 'schemas')

# Función para cargar un schema desde un archivo JSON
def cargar_schema(ruta_archivo):
    with open(ruta_archivo, 'r', encoding='utf-8') as f:
        # Remover comentarios del JSON
        content = f.read()
        lines = content.split('\n')
        clean_lines = [line.split('//')[0] for line in lines]
        clean_content = '\n'.join(clean_lines)
        return json.loads(clean_content)

# Función para cargar todos los schemas de la carpeta
def cargar_todos_los_schemas(directorio):
    schemas = []
    # Buscar todos los archivos JSON en el directorio
    patron = os.path.join(directorio, '*.json')
    archivos_schema = glob.glob(patron)
    
    for archivo in archivos_schema:
        try:
            schema = cargar_schema(archivo)
            # Validar que el schema tenga los campos requeridos
            if 'coleccion' in schema and 'tipo' in schema and 'estructuraNombreProducto' in schema:
                schemas.append(schema)
                print(f'Schema cargado: {os.path.basename(archivo)} (Tipo: {schema["tipo"]})')
            else:
                print(f'Schema inválido: {os.path.basename(archivo)} - Faltan campos requeridos')
        except Exception as e:
            print(f'Error al cargar {os.path.basename(archivo)}: {e}')
    
    return schemas

# Función para convertir plural a singular en español
def plural_a_singular(palabra):
    # Convertir una palabra de plural a singular
    palabra_lower = palabra.lower()
    
    # Reglas de pluralización en español (aplicadas en reversa)
    # Palabras que terminan en -ces → -z (ej: luces → luz)
    if palabra_lower.endswith('ces'):
        base = palabra[:-3]
        return base + 'z'
    
    # Palabras que terminan en -es después de consonante (ej: cables → cable)
    elif palabra_lower.endswith('es') and len(palabra) > 2:
        # Verificar si antes de 'es' hay una consonante
        if palabra[-3] not in 'aeiouáéíóú':
            return palabra[:-2]
        # Si termina en -ies, cambiar a -y (aunque es menos común en español)
        elif palabra_lower.endswith('ies'):
            return palabra[:-3] + 'y'
        # Para palabras que terminan en vocal + es, quitar solo la 's'
        else:
            return palabra[:-1]
    
    # Palabras que terminan en -s después de vocal (ej: laptops → laptop, autos → auto)
    elif palabra_lower.endswith('s') and len(palabra) > 1:
        # Verificar si antes de 's' hay una vocal
        if palabra[-2] in 'aeiouáéíóú':
            return palabra[:-1]
        # Si termina en consonante + s, generalmente quitar solo la 's'
        else:
            return palabra[:-1]
    
    # Si no coincide con ninguna regla, devolver la palabra original
    return palabra

# Función para convertir texto de plural a singular
def texto_a_singular(texto):
    if not texto:
        return texto
    
    # Dividir el texto en palabras
    palabras = texto.split()
    palabras_singulares = []
    
    for palabra in palabras:
        # Preservar mayúsculas/minúsculas originales
        if palabra.isupper():
            # Si toda la palabra está en mayúsculas
            palabras_singulares.append(plural_a_singular(palabra).upper())
        elif palabra[0].isupper():
            # Si solo la primera letra está en mayúscula
            singular = plural_a_singular(palabra)
            palabras_singulares.append(singular[0].upper() + singular[1:])
        else:
            # Minúsculas
            palabras_singulares.append(plural_a_singular(palabra))
    
    return ' '.join(palabras_singulares)

# Función para aplicar transformación de texto
def aplicar_transformacion(texto, transformacion):
    if not texto:
        return None
    
    if transformacion == 'mayuscula':
        return texto.upper()
    elif transformacion == 'minuscula':
        return texto.lower()
    elif transformacion == 'capitalize':
        return texto.capitalize()
    elif transformacion == 'singular':
        return texto_a_singular(texto)
    else:  # 'ninguna'
        return texto

# Función para extraer valor de especificaciones
def extraer_especificacion(especificaciones, titulo_seccion, dato):
    for seccion in especificaciones:
        if seccion.get('tituloSeccion') == titulo_seccion:
            for item in seccion.get('seccionList', []):
                dato_item = item.get('dato')
                # Si dato es una lista, buscar el primer match
                if isinstance(dato, list):
                    if dato_item in dato:
                        return item.get('valor')
                # Si dato es un string, comparar directamente
                elif dato_item == dato:
                    return item.get('valor')
    return None

# Función para procesar un campo del schema
def procesar_campo(producto, campo_config):
    # Si es solo texto estático
    if 'texto' in campo_config:
        return {'valor': campo_config['texto'], 'faltante': False, 'esTexto': True}
    
    valor = None
    
    # Procesar según el tipo de campo
    if campo_config['campo'] == 'marca':
        valor = producto.get('marca')
    
    elif campo_config['campo'] == 'nombreProducto':
        valor = producto.get('nombreProducto')
    
    elif campo_config['campo'] == 'categorias':
        index = campo_config.get('index', 0)
        categorias = producto.get('categorias', [])
        if categorias:
            categoria = categorias[index] if index != -1 else categorias[-1]
            if categoria and 'subcampo' in campo_config:
                valor = categoria.get(campo_config['subcampo'])
    
    elif campo_config['campo'] == 'especificaciones' and 'condicion' in campo_config:
        titulo_seccion = campo_config['condicion']['tituloSeccion']
        # Extraer el dato, que puede ser string o lista
        try:
            dato = campo_config['subcampo']['seccionList']['condicion']['dato']
            valor = extraer_especificacion(producto.get('especificaciones', []), titulo_seccion, dato)
        except (KeyError, TypeError) as e:
            # Si hay error al acceder a la estructura, devolver None
            valor = None
    
    # Aplicar transformación si existe
    if valor and 'transformacion' in campo_config:
        valor = aplicar_transformacion(valor, campo_config['transformacion'])
    
    return {
        'valor': valor,
        'faltante': not valor,
        'esTexto': False
    }

# Función para obtener el nombre del campo para el encabezado
def obtener_nombre_campo(campo_config):
    if campo_config['campo'] == 'especificaciones' and 'condicion' in campo_config:
        # Para especificaciones, usar el nombre del dato específico
        try:
            dato = campo_config['subcampo']['seccionList']['condicion']['dato']
            # Si dato es una lista, usar el primero o concatenar
            if isinstance(dato, list):
                return dato[0] if dato else 'especificaciones'
            return dato
        except (KeyError, TypeError):
            return 'especificaciones'
    else:
        # Para otros campos, usar el nombre del campo
        return campo_config['campo']

# Función para generar el nombre del producto
def generar_nombre_producto(producto, estructura):
    partes = []
    partes_detalle = []
    algun_faltante = False
    
    for campo_config in estructura:
        resultado = procesar_campo(producto, campo_config)
        
        # Para el nombre completo
        if resultado['faltante'] and 'campo' in campo_config:
            partes.append(f"[{campo_config['campo'].upper()}]")
            algun_faltante = True
        elif resultado['valor']:
            partes.append(resultado['valor'])
        
        # Para las columnas individuales (solo campos, no texto estático)
        if 'campo' in campo_config:
            nombre_campo = obtener_nombre_campo(campo_config)
            partes_detalle.append({
                'nombre': nombre_campo,
                'valor': f"[{campo_config['campo'].upper()}]" if resultado['faltante'] else resultado['valor'],
                'faltante': resultado['faltante']
            })
    
    return {
        'nombreCompleto': ' '.join(partes).strip(),
        'partes': partes_detalle,
        'algunFaltante': algun_faltante
    }

# Función para generar una pestaña de Excel para un schema específico
def generar_pestana_excel(ws, productos, estructura, nombre_pestana):
    ws.title = nombre_pestana
    
    # Definir estilos
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    verde_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    amarillo_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    celda_faltante_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
    celda_faltante_font = Font(bold=True)
    
    # Preparar encabezados
    headers = ['SKU', 'Nombre Completo']
    for campo in estructura:
        if 'campo' in campo:
            nombre_campo = obtener_nombre_campo(campo)
            headers.append(nombre_campo)
    
    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Escribir datos
    for row_num, producto in enumerate(productos, 2):
        resultado = generar_nombre_producto(producto, estructura)
        
        # Determinar color de fila
        row_fill = amarillo_fill if resultado['algunFaltante'] else verde_fill
        
        # SKU
        cell = ws.cell(row=row_num, column=1)
        cell.value = producto.get('sku', '')
        cell.fill = row_fill
        
        # Nombre completo
        cell = ws.cell(row=row_num, column=2)
        cell.value = resultado['nombreCompleto']
        cell.fill = row_fill
        
        # Partes individuales
        for col_num, parte in enumerate(resultado['partes'], 3):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = parte['valor'] or ''
            
            # Si la celda tiene dato faltante, aplicar estilo especial
            if parte['faltante']:
                cell.fill = celda_faltante_fill
                cell.font = celda_faltante_font
            else:
                cell.fill = row_fill
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

# Función para generar Excel con múltiples pestañas
def generar_excel_multi_pestana(datos_schemas, nombre_archivo):
    # Crear workbook
    wb = openpyxl.Workbook()
    # Eliminar la pestaña por defecto
    wb.remove(wb.active)
    
    # Crear una pestaña por cada schema
    for schema_data in datos_schemas:
        productos = schema_data['productos']
        estructura = schema_data['estructura']
        nombre_pestana = schema_data['nombre_pestana']
        
        if productos:  # Solo crear pestaña si hay productos
            ws = wb.create_sheet(title=nombre_pestana)
            generar_pestana_excel(ws, productos, estructura, nombre_pestana)
    
    # Guardar archivo
    wb.save(nombre_archivo)
    return len([d for d in datos_schemas if d['productos']])

# Función para validar y limpiar el nombre de pestaña de Excel
def validar_nombre_pestana(nombre):
    # Excel tiene límites: máximo 31 caracteres, no puede contener: \ / ? * [ ]
    caracteres_prohibidos = ['\\', '/', '?', '*', '[', ']']
    nombre_limpio = nombre
    
    # Remover caracteres prohibidos
    for char in caracteres_prohibidos:
        nombre_limpio = nombre_limpio.replace(char, '')
    
    # Limitar a 31 caracteres
    if len(nombre_limpio) > 31:
        nombre_limpio = nombre_limpio[:31]
    
    # Si quedó vacío, usar un nombre por defecto
    if not nombre_limpio:
        nombre_limpio = "Productos"
    
    return nombre_limpio

# Función principal
def main():
    try:
        print('=' * 60)
        print('Generador de Nombres de Productos - Múltiples Schemas')
        print('=' * 60)
        
        # Cargar todos los schemas
        print(f'\nCargando schemas desde: {SCHEMAS_DIR}')
        schemas = cargar_todos_los_schemas(SCHEMAS_DIR)
        
        if not schemas:
            print('\nNo se encontraron schemas válidos para procesar')
            return
        
        print(f'\nTotal de schemas cargados: {len(schemas)}')
        
        # Conectar a MongoDB
        print('\nConectando a MongoDB...')
        client = MongoClient(MONGO_URI)
        print('Conectado exitosamente')
        
        db = client[DB_NAME]
        
        # Procesar cada schema
        datos_schemas = []
        total_productos = 0
        
        print('\n' + '-' * 60)
        print('Procesando schemas...')
        print('-' * 60)
        
        for schema in schemas:
            try:
                tipo = schema['tipo']
                coleccion_nombre = schema['coleccion']
                estructura = schema['estructuraNombreProducto']
                
                print(f'\nProcesando: {tipo}')
                print(f'   Colección: {coleccion_nombre}')
                
                # Obtener productos de la colección correspondiente
                collection = db[coleccion_nombre]
                query = {'categorias.clave': tipo}
                productos = list(collection.find(query))
                
                print(f'Productos encontrados: {len(productos)}')
                
                # Validar nombre de pestaña
                nombre_pestana = validar_nombre_pestana(tipo)
                
                # Agregar a la lista de datos
                datos_schemas.append({
                    'tipo': tipo,
                    'nombre_pestana': nombre_pestana,
                    'productos': productos,
                    'estructura': estructura
                })
                
                total_productos += len(productos)
                
            except Exception as e:
                print(f'Error al procesar {tipo}: {e}')
                continue
        
        print('\n' + '-' * 60)
        print(f'Procesamiento completado')
        print(f'  Total de schemas procesados: {len(datos_schemas)}')
        print(f'  Total de productos: {total_productos}')
        print('-' * 60)
        
        # Generar Excel con múltiples pestañas
        if datos_schemas:
            nombre_archivo = 'productos_output.xlsx'
            print(f'\nGenerando Excel: {nombre_archivo}')
            pestanas_creadas = generar_excel_multi_pestana(datos_schemas, nombre_archivo)
            print(f'Excel generado exitosamente')
            print(f'  Pestañas creadas: {pestanas_creadas}')
            print(f'  Total de productos procesados: {total_productos}')
        else:
            print('\nNo se encontraron productos para generar el Excel')
        
    except Exception as error:
        print(f'\nError: {error}')
        import traceback
        traceback.print_exc()
    finally:
        if 'client' in locals():
            client.close()
            print('\nConexión cerrada')

# Ejecutar
if __name__ == '__main__':
    main()